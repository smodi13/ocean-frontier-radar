"""ARMADA forward-requirements scenario model (Phase 3G).

NOT a financial projection. ARMADA has no publicly known revenue, so there is
no history to project and none is invented. This is a REQUIREMENTS model: it
asks what commercial scale would have to be reached for a venture outcome, and
what unit economics would make that plausible.

Every input is labelled:
    OBSERVED  - traceable to a public source (cited in the workbook)
    ASSUMPTION - analyst input, not sourced

The workbook is written with live Excel formulas, not hardcoded values, so a
reader can change an assumption and see the model recompute.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from ofr import db

OUT = db.ROOT / "models" / "armada_underwriting.xlsx"

HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
OBS = PatternFill("solid", fgColor="E2EFDA")   # observed = green
ASM = PatternFill("solid", fgColor="FFF2CC")   # assumption = amber
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# ---------------------------------------------------------------- inputs
# (label, bear, base, bull, kind, note)
ASSUMPTIONS = [
    ("EPADS pod ASP ($)", 12000, 25000, 45000, "ASSUMPTION",
     "No public price exists. Anchored on the $5kg A-size module form factor and "
     "defence subsystem norms. The single most sensitive input."),
    ("EPADS pods sold per year (steady state)", 150, 600, 2500, "ASSUMPTION",
     "Pods are consumed per deployment if left on the seafloor with the payload. "
     "Volume is the core uncertainty."),
    ("EPADS gross margin", 0.35, 0.50, 0.60, "ASSUMPTION",
     "Hardware consumable; margin depends on contract manufacturing."),
    ("Propulsion module ASP ($)", 20000, 40000, 70000, "ASSUMPTION",
     "Durable subsystem, one per vehicle."),
    ("Propulsion modules per year", 20, 120, 500, "ASSUMPTION",
     "Requires an OEM design-in. Zero such relationship is public today."),
    ("Propulsion gross margin", 0.40, 0.55, 0.65, "ASSUMPTION",
     "Higher than pods; lower volume."),
    ("Engineering / R&D contract revenue ($/yr)", 1500000, 2000000, 2500000, "OBSERVED-anchored",
     "Anchored on observed run-rate: $2,972,287 of federal awards over ~5 years, "
     "with $499,949 obligated in Mar 2026."),
    ("Engineering gross margin", 0.15, 0.25, 0.35, "ASSUMPTION",
     "Cost-plus style R&D work carries thin margin."),
    ("Support / service revenue ($/yr)", 0, 250000, 900000, "ASSUMPTION",
     "Speculative; no evidence of a service line today."),
    ("Headcount at steady state", 12, 25, 45, "ASSUMPTION",
     "Three people are publicly identified today."),
    ("Fully loaded cost per head ($)", 190000, 210000, 230000, "ASSUMPTION",
     "Massachusetts marine/defence engineering."),
    ("Non-headcount opex ($/yr)", 400000, 700000, 1200000, "ASSUMPTION",
     "Facilities, test time, vessel time, certification, IP."),
    ("Years to steady state", 7, 5, 4, "ASSUMPTION",
     "Defence qualification cycles are long."),
    ("Venture-scale revenue threshold ($)", 30000000, 30000000, 30000000, "ASSUMPTION",
     "A common proxy for a company capable of returning a small early-stage fund."),
    ("Exit revenue multiple (x)", 3.0, 5.0, 8.0, "ASSUMPTION",
     "Defence-adjacent hardware trades below software multiples."),
]

OBSERVED_FACTS = [
    ("Total verified federal awards", 2972287, "USAspending + SBIR bulk, 6 distinct awards"),
    ("Navy Phase II obligated (N68335-23-C-0142)", 1998926, "base $999,028 + option $499,949 + CLIN0004 $499,949"),
    ("Most recent obligation", 499949, "Mod P00003, 2026-03-10, incrementally funding CLIN 0004"),
    ("Narrow addressable procurement observed (13 yrs)", 10740202, "components/spares + payload deployment"),
    ("Narrow addressable, annualised", 826169, "the same figure divided by the 13-year span"),
    ("Broad adjacency procurement, annualised", 6272080, "incl. platforms, sensor payloads, launch & recovery"),
    ("Publicly identified employees", 3, "armadamarinerobotics.com/team"),
    ("Known commercial revenue", 0, "no commercial sale evidenced in any public source"),
]


def _label(ws, row, text, fill=HDR, font=WHITE, span=8):
    ws.cell(row=row, column=1, value=text).fill = fill
    ws.cell(row=row, column=1).font = font
    for c in range(2, span + 1):
        ws.cell(row=row, column=c).fill = fill


def build(path: Path = OUT) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ---------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    lines = [
        ("ARMADA Marine Robotics — Forward Requirements / Scenario Model", True),
        ("", False),
        ("WHAT THIS IS", True),
        ("A requirements model, not a projection. ARMADA has no publicly known revenue,", False),
        ("so none is invented and no revenue history is shown. The model asks:", False),
        ("   1. What commercial scale would ARMADA need for a venture-scale outcome?", False),
        ("   2. What unit economics would make that plausible?", False),
        ("", False),
        ("WHAT THIS IS NOT", True),
        ("Not a valuation. Not a forecast. Not an investment recommendation.", False),
        ("Government SBIR/contract obligations are NOT treated as product revenue.", False),
        ("", False),
        ("COLOUR KEY", True),
        ("  Green  = OBSERVED, traceable to a cited public source", False),
        ("  Amber  = ANALYST ASSUMPTION, not sourced. Change these.", False),
        ("", False),
        ("All outputs are live Excel formulas referencing the Assumptions sheet.", False),
        ("Regenerate with: python3 src/ofr/models/armada_underwriting.py", False),
        (f"Generated {db.today()} from public sources accessed 2026-08-21.", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        if bold:
            c.font = BOLD
    ws.column_dimensions["A"].width = 96
    ws.cell(row=14, column=2).fill = OBS
    ws.cell(row=15, column=2).fill = ASM

    # ---------------------------------------------------- Observed facts
    ws = wb.create_sheet("Observed_Facts")
    _label(ws, 1, "OBSERVED FACTS — every value traceable to a public source", span=4)
    for j, h in enumerate(["Fact", "Value", "Source / basis"], start=1):
        c = ws.cell(row=2, column=j, value=h); c.font = BOLD; c.fill = SUB
    for i, (name, val, src) in enumerate(OBSERVED_FACTS, start=3):
        ws.cell(row=i, column=1, value=name)
        vc = ws.cell(row=i, column=2, value=val); vc.fill = OBS
        vc.number_format = '#,##0'
        ws.cell(row=i, column=3, value=src)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 74

    # ------------------------------------------------------- Assumptions
    ws = wb.create_sheet("Assumptions")
    _label(ws, 1, "ASSUMPTIONS — amber cells are ANALYST INPUTS, not sourced facts", span=6)
    for j, h in enumerate(["Driver", "Bear", "Base", "Bull", "Type", "Note"], start=1):
        c = ws.cell(row=2, column=j, value=h); c.font = BOLD; c.fill = SUB
    for i, (name, bear, base, bull, kind, note) in enumerate(ASSUMPTIONS, start=3):
        ws.cell(row=i, column=1, value=name)
        for col, val in ((2, bear), (3, base), (4, bull)):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = OBS if kind.startswith("OBSERVED") else ASM
            c.number_format = '0.0%' if isinstance(val, float) and val < 1 else '#,##0'
            c.border = THIN
        ws.cell(row=i, column=5, value=kind)
        ws.cell(row=i, column=6, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 42
    for col in "BCD":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 82

    # ------------------------------------------------------- Scenarios
    ws = wb.create_sheet("Scenarios")
    _label(ws, 1, "SCENARIO OUTPUTS — live formulas referencing 'Assumptions'", span=4)
    R = {name: i for i, (name, *_rest) in enumerate(ASSUMPTIONS, start=3)}
    cols = {"Bear": "B", "Base": "C", "Bull": "D"}

    rows = [
        ("EPADS revenue", "={c}{p}*{c}{q}", ["EPADS pod ASP ($)", "EPADS pods sold per year (steady state)"]),
        ("Propulsion revenue", "={c}{p}*{c}{q}", ["Propulsion module ASP ($)", "Propulsion modules per year"]),
        ("Engineering revenue", "={c}{p}", ["Engineering / R&D contract revenue ($/yr)"]),
        ("Support revenue", "={c}{p}", ["Support / service revenue ($/yr)"]),
    ]
    ws.cell(row=2, column=1, value="Steady-state revenue build").font = BOLD
    for j, (name, col) in enumerate(cols.items(), start=2):
        c = ws.cell(row=2, column=j, value=name); c.font = BOLD; c.fill = SUB

    r = 3
    line_rows = {}
    for label, tmpl, deps in rows:
        ws.cell(row=r, column=1, value=label)
        for j, (scen, col) in enumerate(cols.items(), start=2):
            if len(deps) == 2:
                f = f"={col}{R[deps[0]]}*{col}{R[deps[1]]}"
            else:
                f = f"={col}{R[deps[0]]}"
            f = f.replace("=", "=Assumptions!").replace("*", "*Assumptions!")
            cell = ws.cell(row=r, column=j, value=f)
            cell.number_format = '#,##0'
        line_rows[label] = r
        r += 1

    tot_rev = r
    ws.cell(row=r, column=1, value="Total revenue").font = BOLD
    for j, col in enumerate(cols.values(), start=2):
        L = get_column_letter(j)
        c = ws.cell(row=r, column=j, value=f"=SUM({L}3:{L}{r-1})")
        c.font = BOLD; c.number_format = '#,##0'
    r += 2

    ws.cell(row=r, column=1, value="Gross profit").font = BOLD
    gp = r
    for j, col in enumerate(cols.values(), start=2):
        L = get_column_letter(j)
        # Support revenue margin is folded in below; kept explicit for clarity.
        f = (f"={L}{line_rows['EPADS revenue']}*Assumptions!{col}{R['EPADS gross margin']}"
             f"+{L}{line_rows['Propulsion revenue']}*Assumptions!{col}{R['Propulsion gross margin']}"
             f"+{L}{line_rows['Engineering revenue']}*Assumptions!{col}{R['Engineering gross margin']}")
        c = ws.cell(row=r, column=j, value=f); c.number_format = '#,##0'; c.font = BOLD
    r += 1

    ws.cell(row=r, column=1, value="Operating cost")
    opex = r
    for j, col in enumerate(cols.values(), start=2):
        f = (f"=Assumptions!{col}{R['Headcount at steady state']}"
             f"*Assumptions!{col}{R['Fully loaded cost per head ($)']}"
             f"+Assumptions!{col}{R['Non-headcount opex ($/yr)']}")
        c = ws.cell(row=r, column=j, value=f); c.number_format = '#,##0'
    r += 1

    ws.cell(row=r, column=1, value="Operating profit").font = BOLD
    for j in range(2, 5):
        L = get_column_letter(j)
        c = ws.cell(row=r, column=j, value=f"={L}{gp}-{L}{opex}")
        c.number_format = '#,##0'; c.font = BOLD
    r += 2

    ws.cell(row=r, column=1, value="Venture-scale gap").font = BOLD
    for j, col in enumerate(cols.values(), start=2):
        L = get_column_letter(j)
        c = ws.cell(row=r, column=j,
                    value=f"=Assumptions!{col}{R['Venture-scale revenue threshold ($)']}-{L}{tot_rev}")
        c.number_format = '#,##0'
    gap = r
    r += 1
    ws.cell(row=r, column=1, value="Reaches venture scale?")
    for j in range(2, 5):
        L = get_column_letter(j)
        ws.cell(row=r, column=j, value=f'=IF({L}{gap}<=0,"YES","NO")')
    r += 1
    ws.cell(row=r, column=1, value="EPADS pods/yr required to close gap alone")
    for j, col in enumerate(cols.values(), start=2):
        L = get_column_letter(j)
        ws.cell(row=r, column=j,
                value=f"=IF({L}{gap}<=0,0,{L}{gap}/Assumptions!{col}{R['EPADS pod ASP ($)']})"
                ).number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="Implied exit value at revenue multiple")
    for j, col in enumerate(cols.values(), start=2):
        L = get_column_letter(j)
        ws.cell(row=r, column=j,
                value=f"={L}{tot_rev}*Assumptions!{col}{R['Exit revenue multiple (x)']}"
                ).number_format = '#,##0'

    ws.column_dimensions["A"].width = 44
    for col in "BCD":
        ws.column_dimensions[col].width = 18

    # ------------------------------------------------ Requirements view
    ws = wb.create_sheet("What_Must_Be_True")
    _label(ws, 1, "WHAT MUST BE TRUE FOR A VENTURE-SCALE OUTCOME", span=3)
    reqs = [
        ("Revenue threshold used", "=Assumptions!C" + str(R["Venture-scale revenue threshold ($)"])),
        ("Pods/yr at Base ASP to hit threshold on EPADS alone",
         f"=Assumptions!C{R['Venture-scale revenue threshold ($)']}/Assumptions!C{R['EPADS pod ASP ($)']}"),
        ("Propulsion modules/yr at Base ASP to hit threshold alone",
         f"=Assumptions!C{R['Venture-scale revenue threshold ($)']}/Assumptions!C{R['Propulsion module ASP ($)']}"),
        ("Observed narrow addressable procurement (annualised)", "=Observed_Facts!B7"),
        ("Threshold as a multiple of observed narrow addressable",
         f"=Assumptions!C{R['Venture-scale revenue threshold ($)']}/Observed_Facts!B7"),
        ("Observed broad adjacency (annualised)", "=Observed_Facts!B8"),
        ("Threshold as a multiple of observed broad adjacency",
         f"=Assumptions!C{R['Venture-scale revenue threshold ($)']}/Observed_Facts!B8"),
    ]
    for i, (lab, formula) in enumerate(reqs, start=3):
        ws.cell(row=i, column=1, value=lab)
        ws.cell(row=i, column=2, value=formula).number_format = '#,##0.0'
    ws.cell(row=len(reqs) + 5, column=1,
            value=("READ THIS: if the threshold is a large multiple of observed addressable "
                   "procurement, the venture case CANNOT rest on the federal market visible "
                   "in our sample. It requires OEM channel, commercial/offshore buyers, "
                   "allied export, or procurement not captured by our keywords.")
            ).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 22

    wb.save(path)
    return path


def compute_python_check() -> dict:
    """Recompute Base case in Python so tests can verify the workbook logic."""
    a = {name: (bear, base, bull) for name, bear, base, bull, *_ in ASSUMPTIONS}
    i = 1  # base index
    epads = a["EPADS pod ASP ($)"][i] * a["EPADS pods sold per year (steady state)"][i]
    prop = a["Propulsion module ASP ($)"][i] * a["Propulsion modules per year"][i]
    eng = a["Engineering / R&D contract revenue ($/yr)"][i]
    sup = a["Support / service revenue ($/yr)"][i]
    total = epads + prop + eng + sup
    gp = (epads * a["EPADS gross margin"][i] + prop * a["Propulsion gross margin"][i]
          + eng * a["Engineering gross margin"][i])
    opex = (a["Headcount at steady state"][i] * a["Fully loaded cost per head ($)"][i]
            + a["Non-headcount opex ($/yr)"][i])
    thr = a["Venture-scale revenue threshold ($)"][i]
    return {"epads": epads, "propulsion": prop, "engineering": eng, "support": sup,
            "total_revenue": total, "gross_profit": gp, "opex": opex,
            "operating_profit": gp - opex, "threshold": thr, "gap": thr - total,
            "reaches_scale": total >= thr}


if __name__ == "__main__":
    p = build()
    r = compute_python_check()
    print(f"workbook -> {p}")
    print("BASE case check:")
    for k, v in r.items():
        print(f"  {k:18s} {v:,.0f}" if isinstance(v, (int, float)) else f"  {k:18s} {v}")
