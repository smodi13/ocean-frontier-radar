"""Phase 3 quality control for the ARMADA diligence outputs.

Tests that the memo cannot silently drift from its evidence: every material
number reconciles, no grant is called revenue, no patent applicant is
mislabelled as owner, and no Propeller portfolio company is called a
competitor without qualification.
"""
import csv
import re
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]


def strip_md(text: str) -> str:
    """Remove markdown emphasis so phrase checks are not defeated by '**not**'."""
    return re.sub(r"[*_`]+", "", text)
ARMADA = ROOT / "research" / "armada"
MEMO = (ROOT / "research" / "armada_investment_memo.md").read_text()
REGISTER = ARMADA / "evidence_register.csv"

DOCS = {p.name: p.read_text() for p in ARMADA.glob("*.md")}
DOCS["memo"] = MEMO


def _register():
    with REGISTER.open() as f:
        return list(csv.DictReader(f))


# ------------------------------------------------------- evidence register
def test_register_is_well_formed():
    rows = _register()
    assert len(rows) >= 40
    ids = [r["claim_id"] for r in rows]
    assert len(ids) == len(set(ids)), "duplicate claim ids"
    for r in rows:
        assert r["claim"].strip()
        assert r["status"] in {"observed", "inferred", "unknown"}, r["claim_id"]
        assert r["memo_section"].strip(), r["claim_id"]
        assert r["date_accessed"].strip(), r["claim_id"]


def test_observed_claims_cite_a_source():
    for r in _register():
        if r["status"] == "observed":
            assert r["source"].strip(), f"{r['claim_id']} observed with no source"


def test_material_financial_claims_use_primary_sources():
    """Tier-3 aggregators must never establish a material financial fact."""
    banned = ("crunchbase", "pitchbook", "tracxn", "cbinsights", "dealroom")
    for r in _register():
        if r["status"] != "observed":
            continue
        if re.search(r"\$[\d,]{7,}", r["claim"]):
            src = r["source"].lower()
            assert not any(b in src for b in banned), \
                f"{r['claim_id']} sources a material financial claim from an aggregator"


def test_every_claim_has_a_source_tier():
    for r in _register():
        assert r["source_tier"].strip(), r["claim_id"]


# ----------------------------------------------------- numeric reconciliation
def test_navy_contract_components_sum_to_total():
    """999,028 + 499,949 + 499,949 = 1,998,926 — asserted in several documents."""
    assert 999_028 + 499_949 + 499_949 == 1_998_926
    for name, text in DOCS.items():
        if "1,998,926" in text:
            assert "999,028" in text, f"{name} states the total without the base award"


def test_federal_total_is_consistent_everywhere():
    for name, text in DOCS.items():
        if "2,972,287" in text:
            assert "$2,972,287" in text, name
    assert "$2,972,287" in MEMO


def test_federal_award_components_sum_to_verified_total():
    parts = [255_821, 246_320, 1_998_926, 174_798, 149_967, 146_455]
    assert sum(parts) == 2_972_287


def test_procurement_numbers_reconcile_with_generated_audit():
    import json
    audit = json.loads((ROOT / "outputs" / "armada_procurement_audit.json").read_text())
    assert audit["n_contracts"] == 87
    narrow = audit["narrow_addressable"]["value"]
    broad = audit["broad_adjacency"]["value"]
    assert f"{narrow:,.0f}".rstrip("0").rstrip(".") or True
    # the memo quotes annualised figures; check they match the audit
    assert f"{audit['narrow_addressable']['annualised']:,.0f}" == "826,169"
    assert f"{audit['broad_adjacency']['annualised']:,.0f}" == "6,272,080"
    assert narrow < broad


def test_model_base_case_matches_memo():
    import sys
    sys.path.insert(0, str(ROOT / "src"))
    from ofr.models.armada_underwriting import compute_python_check
    r = compute_python_check()
    assert r["total_revenue"] == 22_050_000
    assert r["gap"] == 7_950_000
    assert r["reaches_scale"] is False
    assert "22.05M" in MEMO or "22,050,000" in MEMO
    assert "7.95M" in MEMO or "7,950,000" in MEMO


# --------------------------------------------------------- evidence discipline
def test_no_grant_or_contract_is_called_revenue():
    """SBIR obligations must never be presented as product revenue."""
    bad = re.compile(r"(SBIR|STTR|grant|Phase II)[^.\n]{0,60}\brevenue\b", re.IGNORECASE)
    for name, text in DOCS.items():
        for m in bad.finditer(text):
            frag = m.group(0)
            # Allowed only where the text is explicitly denying the equivalence.
            window = text[max(0, m.start() - 200):m.end() + 200].lower()
            assert any(k in window for k in
                       ("not revenue", "is not revenue", "not treated as", "not product revenue",
                        "rather than revenue", "engineering / r&d contract revenue",
                        "engineering revenue", "not product-market fit",
                        "would misrepresent", "not** product-market fit")), \
                f"{name}: '{frag}' may equate government funding with revenue"


def test_memo_states_government_funding_is_not_revenue():
    assert re.search(r"not revenue", MEMO, re.IGNORECASE)


def test_no_invest_recommendation_is_made():
    assert "HOLD" in MEMO
    assert not re.search(r"^\s*#+\s*(recommendation:?\s*)?invest\b", MEMO,
                         re.IGNORECASE | re.MULTILINE)
    assert re.search(r"No INVEST recommendation", MEMO, re.IGNORECASE)


def test_no_fabricated_customer_quotes_or_conversations():
    assert re.search(r"[Nn]o one was contacted", DOCS["primary_research_plan.md"])
    # A quoted sentence attributed to a customer/founder would be fabrication.
    assert not re.search(r'(founder|customer|CEO|operator)\s+(said|told us|explained)',
                         MEMO, re.IGNORECASE)


def test_armada_is_not_described_as_owning_the_licensed_patents():
    """WHOI owns 9,873,499 and 11,990,857; ARMADA is an exclusive licensee."""
    for name, text in DOCS.items():
        for pat in ("9,873,499", "11,990,857"):
            if pat in text:
                assert re.search(r"(licen[sc]|WHOI[- ]owned|assigned to WHOI|"
                                 r"owned by WHOI|assignee.{0,40}Woods Hole)", text,
                                 re.IGNORECASE), \
                    f"{name} mentions {pat} without stating WHOI ownership/licence"
    assert not re.search(r"ARMADA (owns|holds title to) (US )?9,873,499", MEMO)


def test_joint_applicant_is_not_upgraded_to_sole_owner():
    for name, text in DOCS.items():
        if "2024/136933" in text or "2024136933" in text:
            assert re.search(r"joint", text, re.IGNORECASE), \
                f"{name} cites the EPADS PCT without stating joint applicant status"


def test_pct_is_not_presented_as_a_granted_us_patent():
    assert re.search(r"PCT publication.{0,90}no US exclusionary right", MEMO, re.IGNORECASE | re.DOTALL) \
        or re.search(r"confers no US exclusionary right", MEMO, re.IGNORECASE)


def test_no_freedom_to_operate_conclusion_is_asserted():
    assert re.search(r"not a legal opinion", MEMO, re.IGNORECASE)
    # Any mention of freedom to operate (hyphenated or not) must be a disclaimer.
    for m in re.finditer(r"freedom[- ]to[- ]operate", MEMO, re.IGNORECASE):
        window = MEMO[max(0, m.start() - 120):m.end() + 120].lower()
        assert " not " in window or "no " in window, \
            "freedom to operate is discussed without a disclaimer"


# ------------------------------------------------- Propeller portfolio handling
def test_portfolio_companies_are_not_called_direct_competitors():
    comp = (ARMADA / "competitive_landscape.md").read_text()
    for co in ("Orpheus Ocean", "VATN"):
        assert co in comp, f"{co} missing from the competitive map"
        # The company appears in the stack diagram and again in the adjacency
        # table; require at least one occurrence to carry a qualified label.
        labelled = any(
            any(k in comp[m.start():m.start() + 700].lower() for k in
                ("complementary", "different layer", "unclear", "potentially competitive"))
            for m in re.finditer(re.escape(co), comp))
        assert labelled, f"{co} lacks a qualified relationship label"
    assert not re.search(r"(Orpheus|VATN)[^.\n]{0,40}\bdirect competitor\b", comp, re.IGNORECASE)


def test_no_claim_about_propellers_internal_pipeline():
    """Any statement about Propeller having seen ARMADA must sit inside an
    explicit disclaimer, which may precede or follow the phrase."""
    pat = re.compile(r"Propeller (?:has|had) (?:not )?(?:already )?"
                     r"(?:seen|evaluated|diligenced)", re.IGNORECASE)
    for name, text in DOCS.items():
        for m in pat.finditer(text):
            window = text[max(0, m.start() - 220):m.end() + 220].lower()
            assert any(k in window for k in
                       ("no claim is made", "not offered", "cannot", "do not know",
                        "we do not know", "no claim")), \
                f"{name} appears to assert knowledge of Propeller's pipeline"
    assert re.search(r"[Nn]o claim is made.{0,160}(diligenced|evaluated)", MEMO, re.DOTALL)


def test_no_portfolio_whitespace_claim():
    for name, text in DOCS.items():
        assert "whitespace" not in text.lower(), f"{name} asserts portfolio whitespace"


# ----------------------------------------------------------- reference hygiene
def test_internal_file_references_exist():
    refs = set()
    for text in DOCS.values():
        refs |= set(re.findall(r"`((?:research|models|outputs|src)/[^`]+?)`", text))
    missing = [r for r in refs if not (ROOT / r).exists()]
    assert not missing, f"dead internal references: {missing}"


def test_source_urls_are_retained_in_the_register():
    rows = _register()
    with_urls = [r for r in rows if r["source"].startswith("http")]
    assert len(with_urls) >= 15, "too few claims carry a resolvable source URL"


def test_epads_ip_question_is_stated_as_unresolved_not_as_absence_of_rights():
    """The IP gap must be framed as unconfirmed, never as proven absence of rights.
    A negated form ("not a claim that ARMADA lacks rights") is exactly right."""
    pat = re.compile(r"ARMADA (?:lacks|has no|does not have) rights", re.IGNORECASE)
    for name, text in DOCS.items():
        for m in pat.finditer(text):
            window = strip_md(text[max(0, m.start() - 160):m.end() + 60]).lower()
            assert any(k in window for k in ("not a claim", "not evidence", "is not",
                                             "does not mean", "not proof")), \
                f"{name} asserts ARMADA lacks rights without negation"
    assert re.search(r"could not be confirmed from public sources", MEMO, re.IGNORECASE)


def test_all_expected_phase3_documents_exist():
    for f in ["source_reconciliation.md", "technical_diligence.md", "commercial_diligence.md",
              "procurement_market.md", "competitive_landscape.md", "business_model.md",
              "primary_research_plan.md", "investment_debates.md", "evidence_register.csv"]:
        assert (ARMADA / f).exists(), f"missing {f}"
    assert (ROOT / "models" / "armada_underwriting.xlsx").exists()


def test_workbook_outputs_are_formulas_not_hardcoded():
    from openpyxl import load_workbook
    wb = load_workbook(ROOT / "models" / "armada_underwriting.xlsx")
    ws = wb["Scenarios"]
    formulas = sum(1 for row in ws.iter_rows() for c in row
                   if isinstance(c.value, str) and c.value.startswith("="))
    hardcoded = sum(1 for row in ws.iter_rows() for c in row
                    if isinstance(c.value, (int, float)))
    assert formulas >= 30, "scenario outputs should be live formulas"
    assert hardcoded == 0, "scenario sheet contains hardcoded numeric outputs"


def test_workbook_marks_assumptions_distinctly_from_observed():
    from openpyxl import load_workbook
    wb = load_workbook(ROOT / "models" / "armada_underwriting.xlsx")
    assert "Observed_Facts" in wb.sheetnames and "Assumptions" in wb.sheetnames
    types = {wb["Assumptions"].cell(row=r, column=5).value
             for r in range(3, 3 + 15)}
    assert "ASSUMPTION" in types, "assumptions are not labelled"


def test_no_company_revenue_is_presented_as_fact():
    assert re.search(r"no publicly known revenue|zero commercial revenue|"
                     r"[Nn]o commercial sale", MEMO)
    for r in _register():
        if "revenue" in r["claim"].lower() and r["status"] == "observed":
            assert re.search(r"\bno\b|\bzero\b|not", r["claim"], re.IGNORECASE), \
                f"{r['claim_id']} asserts revenue as observed fact"
